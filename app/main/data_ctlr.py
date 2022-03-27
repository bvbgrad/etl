"""
"""

from plistlib import InvalidFileException
import app.utils6L.utils6L as utils

import csv
import logging
import os
import sys

import PySimpleGUI as sg

from datetime import datetime
from dateutil.parser import parse as date_parser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

logger_name = os.getenv("LOGGER_NAME")
logger = logging.getLogger(logger_name)

monthly_filename_remember = ''
monthly_list = []
patient_set = set()
dailyList = set()
used_files = set()


@utils.log_wrap
def xlsx_reader(filename):
    logger.info(__name__ + ".xlsx_reader()")

    rows = []
    try:
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb.active
        for row in ws.values:
            rows.append(row)
        wb.close()
    except Exception:
        sg.popup_error(f"eMARS does not support .csv file format.\nSupported formats are: .xlsx,.xlsm,.xltx,.xltm")
    return rows


@utils.log_wrap
def load_master_data(window):
    logger.info(__name__ + ".load_master_data()")

    monthly_filename = sg.popup_get_file('Select the monthly report')
    if monthly_filename is None or monthly_filename == '':
        sys.exit()

    global monthly_filename_remember
    global patient_set
    monthly_filename_remember = monthly_filename

    print(f"Log intiated at: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    rows_list = xlsx_reader(monthly_filename)
    if len(rows_list) == 0:
        logger.info(f"Master Account file has no data: {monthly_filename}")
        print(f"Master Account file has no data: {monthly_filename}")
        return None

    logger.info(f"Master Account file loaded: {monthly_filename}")
    print(f"Master Account file loaded: {monthly_filename}")
    for i, row in enumerate(rows_list):
        if i == 0:
            logger.info(f'Base column names are {", ".join(row)}')
            continue
    # column names: Patient ID,Patient Name,Billing Code,Duration,Type,Disease Conditions
        patient_set.add(row[1])

        csv_row = {}
        csv_row['Patient Name'] = row[1]
        csv_row['Billing Code'] = row[2]
        csv_row['Duration'] = row[3]
        csv_row['readings'] = 0
        monthly_list.append(csv_row)
    logger.info(f'The Master Report list has {len(monthly_list)} patient records.')
    print(f'The Master Report list has {len(monthly_list)} patient records.')
    
    window['-STATUS-'].update(f"Master Report file has {len(monthly_list)} patient account records.")
    return monthly_filename

@utils.log_wrap
def get_monthly_list():
    logger.info(__name__ + ".get_monthly_list()")
    return monthly_list


@utils.log_wrap
def get_monthly_list_summary():
    logger.info(__name__ + ".get_monthly_list_summary()")

    monthly_list_summary = []
    for i, item in enumerate(monthly_list, 1):
        account_row = (i, item['Patient Name'], item['Billing Code'], item['Duration'], item['readings'])
        monthly_list_summary.append(account_row)
    logger.info(f"The monthly patient summary has {len(monthly_list_summary)} items")
    return monthly_list_summary


@utils.log_wrap
def get_monthly_summary_csv_data():
    logger.info(__name__ + ".get_monthly_summary_csv_data()")

    monthly_list_summary = []
    for item in monthly_list:
        monthly_list_summary.append(item)
    logger.info(f"The monthly account readings summary has {len(monthly_list_summary)} items")
    return monthly_list_summary


@utils.log_wrap
def load_daily_report(window):
    logger.info(__name__ + ".load_daily_report()")    

    daily_filename = sg.popup_get_file('Select a daily report')
    if daily_filename is None or daily_filename == '':
        window['-STATUS-'].update(f"No daily file was selected.")
        return

    if daily_filename in used_files:
        logger.info(f"File '{daily_filename}' has already been used.")
        sg.popup_ok(f"'{daily_filename}'\nhas already been used.")
        window['-STATUS-'].update(f"Loading of {daily_filename} failed")
        window['-REPORT_TEXT_1-'].update(f"File has already been loaded. Please try again.")
        return
    else:
        used_files.add(daily_filename)

    global dailyList
    global monthly_list
    global patient_set

    daily_rows = xlsx_reader(daily_filename)
    if len(daily_rows) == 0:
        logger.info(f"Unable to load daily report '{daily_filename}'")
        return None

    logger.info(f"Loaded daily report '{daily_filename}'")
    print(f"Loaded daily report\n    '{daily_filename}'")
    bDate = False
    count_not_in_master_list = 0
    for i, row in enumerate(daily_rows, 1):
        if row[0] in patient_set:
            patientName = row[0]
    # anticipate the next row will be the patient birth date
            bDate = True
            continue
        if bDate:
    # skip birth date row
            bDate = False
        else:
            try:
                row_date = date_parser(row[0])
                if row[1] in ('MALE', 'FEMALE'):
                    if prior_row0 != '':
                        patientName = prior_row0
                        # print(f"Warning: Patient on line '{i:5}' not in the master account list.")
                        count_not_in_master_list += 1
                    continue
                reading_date = row_date.strftime('%Y-%m-%d')
                dailyList.add((patientName, reading_date,))
            except Exception:
                prior_row0 = row[0]
                continue
    print(f"Warning: There are {count_not_in_master_list} patients that are not in the Master Account list.")
    logger.warn(f"Warning: There are {count_not_in_master_list} patients that are not in the Master Account list.")

    histogram_count = {}
    for patientName, readingDate in dailyList:
        try:
            histogram_count[patientName] += 1
        except KeyError:
            histogram_count[patientName] = 1
    # print(f"readings per Patient ID = {histogram_count}")
    print(f"There are {len(histogram_count)} patients that have readings.")
    logger.warn(f"There are {len(histogram_count)} patients that have readings.")

    # count_zero_readings = set()
    for patientName, new_count in histogram_count.items():
        for i, list_item in enumerate(monthly_list):
            patient = list_item['Patient Name']
            old_count = list_item['readings']
            if patientName == patient:
                monthly_list[i]['readings'] = new_count
                logger.info(f"Updated monthly list record {i:4} from {old_count} to {new_count}")
                # break
            # else:
    #             count_zero_readings.add(patientName)
    # print(f"There are {len(count_zero_readings)} patients that have zero readings.")
    # logger.warn(f"There are {len(count_zero_readings)} patients that have zero readings.")

    window['-STATUS-'].update(f"Summary report updated.")
    window['-REPORT_TEXT_1-'].update('You may continue to add daily reports')
    return Path(daily_filename).name


@utils.log_wrap
def save_report(window):
    logger.info(__name__ + ".save_report()")

    monthly_list_summary = get_monthly_summary_csv_data()
    if len(monthly_list_summary) == 0:
        window['-STATUS-'].update('Summary report is empty. Nothing to save.')
        return

    save_folder = sg.popup_get_folder('Choose folder for saving merged report')
    if save_folder is None or save_folder == '':
        window['-STATUS-'].update("'Save Report' operation canceled.")
        return

    global monthly_filename_remember
    dir_name = os.path.dirname(monthly_filename_remember)
    file_name_csv = os.path.join(dir_name, 'account_summary.csv')
    file_name_xlsx = os.path.join(dir_name, 'account_summary.xlsx')
    fieldnames = ['Patient Name', 'Billing Code', 'Duration', 'readings']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Patient Summary'
    ws.append(fieldnames)
    for row in monthly_list_summary:
        name, code, duration, readings = row.values()
        ws.append((name, code, duration, readings,))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws["A1"].fill = PatternFill("solid", start_color="c9c9c9")
    ws["B1"].fill = PatternFill("solid", start_color="c9c9c9")
    ws["C1"].fill = PatternFill("solid", start_color="c9c9c9")
    ws["D1"].fill = PatternFill("solid", start_color="c9c9c9")

    try:
        wb.save(file_name_xlsx)
        print(f"Summary Excel report saved to '{file_name_xlsx}'.  It has {len(monthly_list_summary)} rows.")

        with open(file_name_csv, mode='w', newline='') as csv_out:
            writer = csv.DictWriter(csv_out, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(monthly_list_summary)
        print(f"Summary CSV report saved to '{file_name_csv}'.  It has {len(monthly_list_summary)} rows.")
        print(f"Summary report saved at: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    except PermissionError:
        sg.popup_error("Could not save 'account_summary' report\nCheck if a previous version is open in Excel.")
        window['-STATUS-'].update("'Save Report' operation canceled.")
        return
    except Exception as err:
        sg.popup_error(f"Error saving account_summary report\n  {err}")
        return
    finally:
        wb.close()

    file_name = os.path.join(dir_name, 'account_summary.log')
    with open(file_name, mode='w') as log_out:
        log_text = window['-LOG-'].get()
        log_out.writelines(log_text)
        logger.info(log_text)
    
    window['-STATUS-'].update(f"'CSV and Log files saved in '{save_folder}'")
    logger.info(f"CSV and Log files saved in '{save_folder}'")
